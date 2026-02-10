from flask import Flask, render_template, request, redirect, jsonify, send_file, session, render_template_string, flash
import json
import os
from datetime import datetime
from fpdf import FPDF
from openpyxl import Workbook

app = Flask(__name__, static_url_path='/static')
app.secret_key = "super_secret_key_123"

PASSWORD = "325979037"

# -----------------------------
#   LOGIN PAGE (HTML)
# -----------------------------
login_page = """
<!DOCTYPE html>
<html lang="he" dir="rtl">
<head>
<meta charset="UTF-8">
<title>התחברות</title>
<style>
    body {
        background: #f0f2f5;
        font-family: Arial, sans-serif;
        display: flex;
        justify-content: center;
        align-items: center;
        height: 100vh;
        margin: 0;
    }
    .login-box {
        background: white;
        padding: 30px;
        border-radius: 12px;
        box-shadow: 0 0 15px rgba(0,0,0,0.1);
        width: 300px;
        text-align: center;
    }
    input {
        width: 90%;
        padding: 10px;
        margin-top: 15px;
        border: 1px solid #ccc;
        border-radius: 8px;
        font-size: 16px;
    }
    button {
        margin-top: 20px;
        width: 100%;
        padding: 12px;
        background: #4a90e2;
        border: none;
        color: white;
        font-size: 16px;
        border-radius: 8px;
        cursor: pointer;
    }
    button:hover {
        background: #357ac9;
    }
</style>
</head>
<body>
    <div class="login-box">
        <h2>התחברות למערכת</h2>
        <form method="POST">
            <input type="password" name="password" placeholder="הכנס סיסמה">
            <button type="submit">כניסה</button>
        </form>
    </div>
</body>
</html>
"""

# -----------------------------
#   LOGIN ROUTE
# -----------------------------
@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        if request.form.get("password") == PASSWORD:
            session["logged_in"] = True
            return redirect("/")
        return "סיסמה שגויה"
    return render_template_string(login_page)

@app.route("/logout")
def logout():
    session.pop("logged_in", None)
    return redirect("/login")

# -----------------------------
#   AUTH CHECK
# -----------------------------
@app.before_request
def require_login():
    allowed = ["login", "logout", "static"]
    if request.endpoint not in allowed and not session.get("logged_in"):
        return redirect("/login")

# -----------------------------
#   LOAD + SAVE DATA
# -----------------------------
DATA_FILE = "data.json"

def load_json():
    if not os.path.exists(DATA_FILE):
        return {"days": [], "active_day": None, "people": []}
    with open(DATA_FILE, "r", encoding="utf-8") as f:
        return json.load(f)

def save_json(data):
    with open(DATA_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

@app.context_processor
def inject_day():
    data = load_json()
    return dict(active_day=data.get("active_day"))

# -----------------------------
#   TOTALS — FIXED FOR day_id
# -----------------------------
def calculate_totals(boys, active_day):
    total_all = 0
    total_daily = 0

    for boy in boys:
        for d in boy["donations"]:
            total_all += int(d["amount"])
            if d.get("day_id") == active_day:
                total_daily += int(d["amount"])

        for d in boy["mobile_donations"]:
            full = int(d.get("full_amount", d["amount"]))
            total_all += full
            if d.get("day_id") == active_day:
                total_daily += full

    return total_daily, total_all

# -----------------------------
#   RESET DATA
# -----------------------------
def reset_data():
    data = load_json()

    # איפוס בחורים
    people = data.get("people", [])
    for boy in people:
        boy["donations"] = []
        boy["mobile_donations"] = []
        boy["status"] = "נמצא"
        boy["place"] = ""

    # איפוס היסטוריית ימים
    data["days"] = []
    data["active_day"] = None

    save_json(data)

# -----------------------------
#   HOME PAGE
# -----------------------------
@app.route("/")
def index():
    data = load_json()
    boys = data.get("people", [])
    active_day = data.get("active_day")

    # חישוב סכומים אמיתיים
    total_daily, total_all = calculate_totals(boys, active_day)

    # סינון (אם יש)
    status = request.args.get("status")

    filtered_boys = boys
    if status:
        filtered_boys = [b for b in boys if b["status"] == status]

    return render_template(
        "index.html",
        boys=filtered_boys,
        total_daily=total_daily,
        total_all=total_all,
        active_day=active_day
    )
# -----------------------------
#   OPEN/CLOSE DAY
# -----------------------------
@app.route("/open_day", methods=["POST"])
def open_day_route():
    data = load_json()

    if data.get("active_day") is not None:
        flash("כבר יש יום פתוח", "danger")
        return redirect("/")

    new_day_id = len(data.get("days", [])) + 1

    new_day = {
        "day_id": new_day_id,
        "opened_at": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "closed_at": None
    }

    data.setdefault("days", []).append(new_day)
    data["active_day"] = new_day_id

    save_json(data)
    flash(f"יום מספר {new_day_id} נפתח", "success")
    return redirect("/")

@app.route("/close_day", methods=["POST"])
def close_day_route():
    password = request.form.get("password")

    # בדיקת סיסמה
    if password != "1234":
        flash("סיסמה שגויה", "danger")
        return redirect("/")

    data = load_json()

    if data.get("active_day") is None:
        flash("אין יום פתוח", "danger")
        return redirect("/")

    for day in data.get("days", []):
        if day["day_id"] == data["active_day"]:
            day["closed_at"] = datetime.now().strftime("%Y-%m-%d %H:%M")

    data["active_day"] = None
    save_json(data)

    flash("היום נסגר", "success")
    return redirect("/")
# -----------------------------
#   FILTER STATUS
# -----------------------------
@app.route("/filter")
def filter_boys():
    data = load_json()
    boys = data.get("people", [])
    active_day = data.get("active_day")

    # ⭐ חישוב סכומים אמיתיים לפני סינון
    total_daily, total_all = calculate_totals(boys, active_day)

    # ⭐ סינון רק להצגת הטבלה
    status = request.args.get("status")
    filtered_boys = boys

    if status:
        filtered_boys = [b for b in boys if b["status"] == status]

    return render_template(
        "index.html",
        boys=filtered_boys,
        total_daily=total_daily,   # תמיד אמיתי
        total_all=total_all,       # תמיד אמיתי
        active_day=active_day      # ⭐ זה היה חסר!
    )
# -----------------------------
#   RESET ROUTE
# -----------------------------
@app.route("/reset", methods=["POST"])
def reset():
    password = request.form.get("password")
    CORRECT_PASSWORD = "0548414987"

    if password != CORRECT_PASSWORD:
        flash("❌ הסיסמה שגויה — האיפוס לא בוצע", "danger")
        return redirect("/")

    reset_data()
    flash("✔️ האיפוס בוצע בהצלחה", "success")
    return redirect("/")

# -----------------------------
#   ADD BOY
# -----------------------------
@app.route("/add_boy", methods=["GET", "POST"])
def add_boy():
    if request.method == "POST":
        name = request.form["name"]
        place = request.form.get("place", "")

        data = load_json()
        people = data.get("people", [])

        new_id = max([b["id"] for b in people], default=0) + 1

        new_boy = {
            "id": new_id,
            "name": name,
            "place": place,
            "status": "נמצא",
            "donations": [],
            "mobile_donations": []
        }

        people.append(new_boy)
        data["people"] = people
        save_json(data)

        return redirect("/")

    return render_template("add_boy.html")

from flask import (
    Flask,
    render_template,
    request,
    redirect,
    jsonify,
    send_file,
    session,
    render_template_string,
    flash,
)
import json
import os
from datetime import datetime
from fpdf import FPDF
from openpyxl import Workbook

app = Flask(__name__, static_url_path="/static")
app.secret_key = "super_secret_key_123"

PASSWORD = "325979037"
DATA_FILE = "data.json"

# -----------------------------
#   LOGIN PAGE (INLINE HTML)
# -----------------------------
login_page = """
<!DOCTYPE html>
<html lang="he" dir="rtl">
<head>
<meta charset="UTF-8">
<title>התחברות</title>
<style>
    body {
        background: #f0f2f5;
        font-family: Arial, sans-serif;
        display: flex;
        justify-content: center;
        align-items: center;
        height: 100vh;
        margin: 0;
    }
    .login-box {
        background: white;
        padding: 30px;
        border-radius: 12px;
        box-shadow: 0 0 15px rgba(0,0,0,0.1);
        width: 300px;
        text-align: center;
    }
    input {
        width: 90%;
        padding: 10px;
        margin-top: 15px;
        border: 1px solid #ccc;
        border-radius: 8px;
        font-size: 16px;
    }
    button {
        margin-top: 20px;
        width: 100%;
        padding: 12px;
        background: #4a90e2;
        border: none;
        color: white;
        font-size: 16px;
        border-radius: 8px;
        cursor: pointer;
    }
    button:hover {
        background: #357ac9;
    }
</style>
</head>
<body>
    <div class="login-box">
        <h2>התחברות למערכת</h2>
        <form method="POST">
            <input type="password" name="password" placeholder="הכנס סיסמה">
            <button type="submit">כניסה</button>
        </form>
    </div>
</body>
</html>
"""

# -----------------------------
#   AUTH / LOGIN
# -----------------------------
@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        if request.form.get("password") == PASSWORD:
            session["logged_in"] = True
            return redirect("/")
        return "סיסמה שגויה"
    return render_template_string(login_page)


@app.route("/logout")
def logout():
    session.pop("logged_in", None)
    return redirect("/login")


@app.before_request
def require_login():
    allowed = ["login", "logout", "static"]
    if request.endpoint not in allowed and not session.get("logged_in"):
        return redirect("/login")


# -----------------------------
#   DATA LOAD / SAVE
# -----------------------------
def load_json():
    if not os.path.exists(DATA_FILE):
        return {"days": [], "active_day": None, "people": []}
    with open(DATA_FILE, "r", encoding="utf-8") as f:
        return json.load(f)


def save_json(data):
    with open(DATA_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


@app.context_processor
def inject_day():
    data = load_json()
    return dict(active_day=data.get("active_day"))


# -----------------------------
#   TOTALS (כולל full_amount)
# -----------------------------
def calculate_totals(boys, active_day):
    total_all = 0
    total_daily = 0

    for boy in boys:
        for d in boy["donations"]:
            total_all += int(d["amount"])
            if d.get("day_id") == active_day:
                total_daily += int(d["amount"])

        for d in boy["mobile_donations"]:
            full = int(d.get("full_amount", d["amount"]))
            total_all += full
            if d.get("day_id") == active_day:
                total_daily += full

    return total_daily, total_all


# -----------------------------
#   RESET DATA
# -----------------------------
def reset_data():
    data = load_json()

    people = data.get("people", [])
    for boy in people:
        boy["donations"] = []
        boy["mobile_donations"] = []
        boy["status"] = "נמצא"
        boy["place"] = ""

    data["days"] = []
    data["active_day"] = None

    save_json(data)


# -----------------------------
#   HOME PAGE
# -----------------------------
@app.route("/")
def index():
    data = load_json()
    boys = data.get("people", [])
    active_day = data.get("active_day")

    # ⭐ חישוב סכומים אמיתיים לפני סינון
    total_daily, total_all = calculate_totals(boys, active_day)

    # ⭐ סינון (רק להצגת הטבלה)
    status_filter = request.args.get("status")
    place_filter = request.args.get("place")

    filtered_boys = boys

    if status_filter:
        filtered_boys = [b for b in filtered_boys if b["status"] == status_filter]

    if place_filter:
        filtered_boys = [b for b in filtered_boys if place_filter in b["place"]]

    # ⭐ מציגים רק את הסכומים האמיתיים — לא מסוננים
    return render_template(
        "index.html",
        boys=filtered_boys,
        total_daily=total_daily,
        total_all=total_all
    )


# -----------------------------
#   OPEN / CLOSE DAY
# -----------------------------
@app.route("/open_day", methods=["POST"])
def open_day_route():
    data = load_json()

    if data.get("active_day") is not None:
        flash("כבר יש יום פתוח", "danger")
        return redirect("/")

    new_day_id = len(data.get("days", [])) + 1

    new_day = {
        "day_id": new_day_id,
        "opened_at": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "closed_at": None
    }

    data.setdefault("days", []).append(new_day)
    data["active_day"] = new_day_id

    save_json(data)
    flash(f"יום מספר {new_day_id} נפתח", "success")
    return redirect("/")


@app.route("/close_day", methods=["POST"])
def close_day_route():
    print(">>> PASSWORD RECEIVED:", request.form.get("password"))

    password = request.form.get("password")

    if password != "1234":
        flash("סיסמה שגויה", "danger")
        return redirect("/")

    data = load_json()

    if data.get("active_day") is None:
        flash("אין יום פתוח", "danger")
        return redirect("/")

    for day in data.get("days", []):
        if day["day_id"] == data["active_day"]:
            day["closed_at"] = datetime.now().strftime("%Y-%m-%d %H:%M")

    data["active_day"] = None
    save_json(data)

    flash("היום נסגר", "success")
    return redirect("/")
# -----------------------------
#   FILTER STATUS
# -----------------------------
@app.route("/filter")
def filter_status():
    status = request.args.get("status", "")
    data = load_json()
    boys = data.get("people", [])

    if status == "":
        filtered = boys
    else:
        filtered = [b for b in boys if b["status"] == status]

    active_day = data.get("active_day")
    total_daily, total_all = calculate_totals(filtered, active_day)

    return render_template(
        "index.html", boys=filtered, total_daily=total_daily, total_all=total_all
    )


# -----------------------------
#   RESET ROUTE
# -----------------------------
@app.route("/reset", methods=["POST"])
def reset():
    password = request.form.get("password")
    CORRECT_PASSWORD = "0548414987"

    if password != CORRECT_PASSWORD:
        flash("❌ הסיסמה שגויה — האיפוס לא בוצע", "danger")
        return redirect("/")

    reset_data()
    flash("✔️ האיפוס בוצע בהצלחה", "success")
    return redirect("/")


# -----------------------------
#   ADD BOY
# -----------------------------
@app.route("/add_boy", methods=["GET", "POST"])
def add_boy():
    if request.method == "POST":
        name = request.form["name"]
        place = request.form.get("place", "")

        data = load_json()
        people = data.get("people", [])

        new_id = max([b["id"] for b in people], default=0) + 1

        new_boy = {
            "id": new_id,
            "name": name,
            "place": place,
            "status": "נמצא",
            "donations": [],
            "mobile_donations": [],
        }

        people.append(new_boy)
        data["people"] = people
        save_json(data)

        return redirect("/")

    return render_template("add_boy.html")


# -----------------------------
#   BOY PAGE
# -----------------------------
@app.route("/boy/<int:boy_id>")
def boy_page(boy_id):
    data = load_json()
    people = data.get("people", [])
    boy = next((b for b in people if b["id"] == boy_id), None)

    if not boy:
        return "לא נמצא", 404

    active_day = data.get("active_day")

    daily = sum(
        int(d["amount"]) for d in boy["donations"] if d.get("day_id") == active_day
    )
    daily += sum(
        int(d["amount"])
        for d in boy["mobile_donations"]
        if d.get("day_id") == active_day
    )

    personal = sum(int(d["amount"]) for d in boy["donations"])
    personal += sum(int(d["amount"]) for d in boy["mobile_donations"])

    return render_template("boy.html", boy=boy, daily=daily, personal=personal)


# -----------------------------
#   UPDATE STATUS
# -----------------------------
@app.route("/update_status/<int:boy_id>", methods=["POST"])
def update_status(boy_id):
    data = load_json()
    people = data.get("people", [])
    boy = next((b for b in people if b["id"] == boy_id), None)

    if not boy:
        return "לא נמצא", 404

    new_status = request.form["status"]
    boy["status"] = new_status

    if new_status == "נמצא":
        boy["place"] = ""

    save_json(data)
    return redirect(f"/boy/{boy_id}")


# -----------------------------
#   ADD DONATION
# -----------------------------
@app.route("/add_donation", methods=["POST"])
def add_donation():
    boy_id = int(request.form["id"])
    amount = request.form["amount"]
    method = request.form["method"]

    data = load_json()
    people = data.get("people", [])
    active_day = data.get("active_day")

    if active_day is None:
        flash("אין יום פתוח — אי אפשר להוסיף תרומות", "danger")
        return redirect(f"/boy/{boy_id}")

    boy = next((b for b in people if b["id"] == boy_id), None)

    if not boy:
        return "לא נמצא", 404

    donation = {
        "amount": amount,
        "date": datetime.now().strftime("%Y-%m-%d"),
        "time": datetime.now().strftime("%H:%M"),
        "day_id": active_day,
    }

    boy[method].append(donation)
    save_json(data)

    return redirect(f"/boy/{boy_id}")


@app.route("/add_mobile_with_percent", methods=["POST"])
def add_mobile_with_percent():
    data = load_json()
    people = data.get("people", [])
    active_day = data.get("active_day")

    req = request.get_json()
    boy_id = int(req["id"])
    amount = int(req["amount"])
    percent = int(req["percent"])

    boy = next((b for b in people if b["id"] == boy_id), None)
    if not boy:
        return "לא נמצא", 404

    amount_for_boy = amount if percent == 100 else amount // 2

    donation = {
        "amount": amount_for_boy,
        "full_amount": amount,
        "percent": percent,
        "date": datetime.now().strftime("%Y-%m-%d"),
        "time": datetime.now().strftime("%H:%M"),
        "day_id": active_day,
    }

    boy["mobile_donations"].append(donation)
    save_json(data)

    return "OK"

@app.route("/check_day_open")
def check_day_open():
    data = load_json()
    active_day = data.get("active_day")
    return {"open": active_day is not None}

# -----------------------------
#   DELETE DONATION
# -----------------------------
@app.route("/delete_donation/<int:boy_id>/<method>/<int:index>")
def delete_donation(boy_id, method, index):
    data = load_json()
    people = data.get("people", [])
    boy = next((b for b in people if b["id"] == boy_id), None)

    if not boy:
        return "לא נמצא", 404

    boy[method].pop(index)
    save_json(data)

    return redirect(f"/boy/{boy_id}")


# -----------------------------
#   EDIT DONATION
# -----------------------------
@app.route("/edit_donation/<int:boy_id>/<method>/<int:index>", methods=["GET", "POST"])
def edit_donation(boy_id, method, index):
    data = load_json()
    people = data.get("people", [])
    boy = next((b for b in people if b["id"] == boy_id), None)

    if not boy:
        return "לא נמצא", 404

    if request.method == "POST":
        new_amount = request.form["amount"]
        boy[method][index]["amount"] = new_amount
        save_json(data)
        return redirect(f"/boy/{boy_id}")

    donation = boy[method][index]
    return render_template(
        "edit_donation.html", boy=boy, donation=donation, method=method, index=index
    )


# -----------------------------
#   TOGGLE STATUS
# -----------------------------
@app.route("/toggle_status/<int:boy_id>", methods=["POST"])
def toggle_status(boy_id):
    data = load_json()
    people = data.get("people", [])
    boy = next((b for b in people if b["id"] == boy_id), None)

    if not boy:
        return "לא נמצא", 404

    boy["status"] = "יצא" if boy["status"] == "נמצא" else "נמצא"
    save_json(data)

    return redirect(f"/boy/{boy_id}")


# -----------------------------
#   SEARCH
# -----------------------------
@app.route("/search")
def search():
    query = request.args.get("q", "").strip()
    data = load_json()
    people = data.get("people", [])

    results = [b for b in people if query in b["name"] or query in b.get("place", "")]
    return render_template("search.html", results=results, query=query)


# -----------------------------
#   UPDATE PLACE
# -----------------------------
@app.route("/update_place/<int:boy_id>", methods=["POST"])
def update_place(boy_id):
    data = load_json()
    people = data.get("people", [])
    boy = next((b for b in people if b["id"] == boy_id), None)

    if not boy:
        return "לא נמצא", 404

    boy["place"] = request.form["place"]
    save_json(data)

    return redirect(f"/boy/{boy_id}")


# -----------------------------
#   REPORTS
# -----------------------------
@app.route("/reports")
def reports():
    data = load_json()
    boys = data.get("people", [])
    return render_template("reports.html", boys=boys)


# -----------------------------
#   DAYS LIST
# -----------------------------
@app.route("/days")
def days_list():
    data = load_json()
    days = data.get("days", [])
    return render_template("days.html", days=days)


# -----------------------------
#   DAY REPORT
# -----------------------------
@app.route("/day/<int:day_id>")
def day_report(day_id):
    data = load_json()
    people = data.get("people", [])

    report = []
    total = 0

    for boy in people:
        boy_total = 0

        for d in boy["donations"]:
            if d.get("day_id") == day_id:
                boy_total += int(d["amount"])

        for d in boy["mobile_donations"]:
            if d.get("day_id") == day_id:
                boy_total += int(d["amount"])

        if boy_total > 0:
            report.append(
                {
                    "id": boy["id"],
                    "name": boy["name"],
                    "total": boy_total,
                }
            )

        total += boy_total

    return render_template(
        "day_report.html", report=report, total=total, day_id=day_id
    )


# -----------------------------
#   EXCEL EXPORT FOR SPECIFIC DAY
# -----------------------------
@app.route("/excel_day/<int:day_id>")
def export_excel_day(day_id):
    data = load_json()
    people = data.get("people", [])

    wb = Workbook()
    ws = wb.active
    ws.title = f"יום {day_id}"

    ws.append(["ID", "שם", "סכום רגיל", "סכום נייד", "סה״כ ליום"])

    for boy in people:
        total_regular = sum(
            int(d["amount"]) for d in boy["donations"] if d.get("day_id") == day_id
        )
        total_mobile = sum(
            int(d["amount"])
            for d in boy["mobile_donations"]
            if d.get("day_id") == day_id
        )
        total_all = total_regular + total_mobile

        if total_all > 0:
            ws.append(
                [
                    boy["id"],
                    boy["name"],
                    total_regular,
                    total_mobile,
                    total_all,
                ]
            )

    filename = f"day_{day_id}.xlsx"
    wb.save(filename)

    return send_file(filename, as_attachment=True)


# -----------------------------
#   PDF EXPORT
# -----------------------------
@app.route("/pdf")
def pdf_export():
    data = load_json()
    people = data.get("people", [])

    pdf = FPDF()
    pdf.add_page()
    pdf.add_font("Arial", "", fname="arial.ttf", uni=True)
    pdf.set_font("Arial", size=14)

    pdf.cell(200, 10, txt="דוח תרומות", ln=True, align="C")

    for boy in people:
        pdf.set_font("Arial", size=12)
        pdf.cell(200, 10, txt=f"{boy['name']}:", ln=True)

        for d in boy["donations"]:
            pdf.cell(
                200,
                8,
                txt=f"  רגיל: {d['amount']} ₪ ({d['date']})",
                ln=True,
            )

        for d in boy["mobile_donations"]:
            pdf.cell(
                200,
                8,
                txt=f"  נייד: {d['amount']} ₪ ({d['date']})",
                ln=True,
            )

        pdf.ln(4)

    pdf.output("report.pdf")
    return send_file("report.pdf", as_attachment=True)


# -----------------------------
#   EXCEL EXPORT (ALL)
# -----------------------------
@app.route("/excel")
def export_excel():
    data = load_json()
    people = data.get("people", [])

    wb = Workbook()
    ws = wb.active
    ws.title = "דוח בחורים"

    ws.append(["ID", "שם", "מקום", "סטטוס", "סכום רגיל", "סכום נייד", "סה״כ"])

    for boy in people:
        total_regular = sum(int(d["amount"]) for d in boy["donations"])
        total_mobile = sum(int(d["amount"]) for d in boy["mobile_donations"])
        total_all = total_regular + total_mobile

        ws.append(
            [
                boy["id"],
                boy["name"],
                boy.get("place", ""),
                boy["status"],
                total_regular,
                total_mobile,
                total_all,
            ]
        )

    filename = "report.xlsx"
    wb.save(filename)
    return send_file(filename, as_attachment=True)


# -----------------------------
#   RUN
# -----------------------------
if __name__ == "__main__":
    app.run(debug=True, host="0.0.0.0", port=80)